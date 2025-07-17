# get_holidays.py (version finale)
import requests
from bs4 import BeautifulSoup
from datetime import datetime
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

def scrape_apchq_holidays():
    """Scrape les congés depuis APCHQ"""
    url = "https://www.apchq.com/nos-services/relations-du-travail/conges/"
    
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
    }
    
    try:
        print("🔍 Récupération des données...")
        response = requests.get(url, headers=headers, timeout=15)
        response.raise_for_status()
        
        soup = BeautifulSoup(response.content, 'html.parser')
        holidays = []
        
        # Chercher les sections spécifiques
        sections = soup.find_all('section', id=['jours-feries-chomes', 'conges-annuels-obligatoires'])
        
        for section in sections:
            holidays.extend(extract_holidays_from_section(section))
        
        return holidays
        
    except Exception as e:
        print(f"❌ Erreur: {e}")
        return []

def extract_holidays_from_section(section):
    """Extrait les congés d'une section"""
    holidays = []
    months_fr = {
        'janvier': 1, 'février': 2, 'mars': 3, 'avril': 4,
        'mai': 5, 'juin': 6, 'juillet': 7, 'août': 8,
        'septembre': 9, 'octobre': 10, 'novembre': 11, 'décembre': 12
    }
    
    lists = section.find_all(['ul', 'ol'])
    
    for ul in lists:
        items = ul.find_all('li')
        for li in items:
            text = li.get_text().strip()
            
            # Pattern : "Nom : date"
            pattern = r'(.+?)\s*:\s*(\d{1,2}(?:er)?)\s+(janvier|février|mars|avril|mai|juin|juillet|août|septembre|octobre|novembre|décembre)\s+(\d{4})'
            match = re.search(pattern, text)
            
            if match:
                name = match.group(1).strip()
                day = int(match.group(2).replace('er', ''))
                month_name = match.group(3)
                year = int(match.group(4))
                
                if month_name in months_fr:
                    month = months_fr[month_name]
                    try:
                        date_obj = datetime(year, month, day)
                        holidays.append({
                            "date": date_obj,
                            "description": name
                        })
                    except ValueError:
                        continue
    
    return holidays

def create_excel_file(holidays, filename="conges_apchq.xlsx"):
    """Crée le fichier Excel"""
    print("📊 Création du fichier Excel...")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Congés APCHQ"
    
    # Headers
    headers = ["Date Début", "Date Fin", "Description"]
    ws.append(headers)
    
    # Style des headers
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Ajouter les données
    for holiday in sorted(holidays, key=lambda x: x["date"]):
        date = holiday["date"]
        ws.append([
            date.strftime("%Y-%m-%dT00:00:00.000Z"),
            date.strftime("%Y-%m-%dT23:59:00.000Z"),
            holiday["description"]
        ])
    
    # Ajuster les colonnes
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 35
    
    wb.save(filename)
    print(f"✅ Fichier créé: {filename}")

def main():
    holidays = scrape_apchq_holidays()
    if holidays:
        create_excel_file(holidays)
        print(f"✅ {len(holidays)} congés trouvés")
    else:
        print("❌ Aucun congé trouvé")

if __name__ == "__main__":
    main()
